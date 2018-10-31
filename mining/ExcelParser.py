
from openpyxl import *
from datetime import datetime, timedelta

def Cryptopia():

   # Load in the workbook
   wb = load_workbook('Deposit_History.xlsx')
   sheet = wb.active

   data = {}
   coin = "ZCL"

   for idx in range(2, sheet.max_row):

      coinName = sheet.cell(row=idx, column=2).value

      if coinName == coin:

         timestamp = sheet.cell(row=idx, column=8).value
         timestamp = timestamp.split(' ')[0].split('/')
         timestamp = "{0}/{1}/{2}".format(timestamp[1], timestamp[0], timestamp[2])

         amount = sheet.cell(row=idx, column=3).value

         total = data.get(timestamp, 0.0)
         total += amount

         data[timestamp] = total

   idx = 2
   dataList = list( data.items() )
   length   = len( dataList ) - 1

   for x in range(length, -1, -1):
      idxDate = 'J{0}'.format(idx)
      idxAmount = 'K{0}'.format(idx)

      sheet[idxDate] = dataList[x][0]
      sheet[idxAmount] = dataList[x][1]

      idx = idx + 1
      # outText = "{0}: {1} on {2}".format(coin, dataList[x][0], dataList[x][1])
      # print( outText )

   # Sava the file
   wb.save('Deposit_History.xlsx')


def Binance():

   # Load in the workbook
   wb = load_workbook('DepositHistory.xlsx')
   sheet = wb.active

   xl_data = {}
   data    = {}
   coins   = set()

   for idx in range(2, sheet.max_row):

      coinName = sheet.cell(row=idx, column=2).value
      coins.add( coinName )
   
   coins = list( coins )
   coins.sort()

   # Add the coins received in a day
   for x in range(0, len(coins) ): 

      for idx in range(2, sheet.max_row):

         coinName = sheet.cell(row=idx, column=2).value

         if coinName == coins[x]:
            
            timestamp = sheet.cell(row=idx, column=1).value
            timestamp = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S") - timedelta(hours=4)
            timestamp = str( timestamp ).split(' ')[0]

            amount = sheet.cell(row=idx, column=3).value

            total = data.get(timestamp, 0.0)
            total += float(amount)

            data[timestamp] = total

            # print(coinName, timestamp, total)

      # Add coin totals to excel dictionary
      xl_data[ coins[x] ] = data
      data = {}

   # Write data to the excel sheet
   for coin in xl_data:

      # Get the coin dictionary 
      dataList = list( xl_data[coin].items() )
      length   = len( dataList ) - 1

      idxTime = 5 + ( coins.index(coin) * 2 )
      idxCoin = 6 + ( coins.index(coin) * 2 )

      sheet.cell(row=1, column=idxTime).value = "Date"
      sheet.cell(row=1, column=idxCoin).value = coin
      
      idx = 2

      for x in range(length, -1, -1):

         sheet.cell(row=idx, column=idxTime).value = dataList[x][0]
         sheet.cell(row=idx, column=idxCoin).value = dataList[x][1]

         idx = idx + 1

         outText = "{0}: {1} on {2}".format(coin, dataList[x][0], dataList[x][1])
         print( outText )

   # Sava the file
   wb.save('DepositHistory.xlsx')


# Cryptopia()
Binance()